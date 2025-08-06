import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import os



def calc_absence_report(input_file, output_file):
    wb_in = openpyxl.load_workbook(input_file)
    ws = wb_in.active

    records = []

    # 生徒名：B10～B36 → 行番号：10～36
    for row in range(10, 37):
        student_name = ws.cell(row=row, column=2).value  # B列＝2
        if not student_name:
            continue

        # C列以降：科目ごとの欠席数
        col = 3
        while True:
            subject = ws.cell(row=9, column=col).value  # 科目名（9行目にある前提）
            if subject is None:
                break  # 科目がない＝終了

            absence_count = ws.cell(row=row, column=col).value
            total_classes = ws.cell(row=2, column=col).value  # 授業数は2行目にある

            # 欠席数や授業数が数値でなければスキップ
            if not isinstance(absence_count, (int, float)):
                absence_count = 0
            if not isinstance(total_classes, (int, float)) or total_classes == 0:
                col += 1
                continue

            absence_rate = absence_count / total_classes

            records.append({
                "student_name": student_name,
                "subject": subject,
                "absence_count": absence_count,
                "total_classes": total_classes,
                "absence_rate": absence_rate
            })

            col += 1



# def calc_absence_report(input_file, output_file):
#     wb_in = openpyxl.load_workbook(input_file)
#     ws_in = wb_in.active

#     # 読み込みデータを辞書形式に変換（ヘッダー前提）
#     records = []
#     for row in ws_in.iter_rows(min_row=2, values_only=True):
#         if len(row) < 4:
#             continue  # 列が足りない行は無視

#         student_name = row[0]
#         subject = row[1]
#         absence_count = row[2]
#         total_classes = row[3]

#     # Noneのときは0に変換
#         absence_count = absence_count if isinstance(absence_count, (int, float)) else 0
#         total_classes = total_classes if isinstance(total_classes, (int, float)) else 0

#     # 授業数が0ならスキップ（0除算防止）
#         if total_classes == 0:
#             continue

#         absence_rate = absence_count / total_classes
#         records.append({
#             "student_name": student_name,
#             "subject": subject,
#             "absence_count": absence_count,
#             "total_classes": total_classes,
#             "absence_rate": absence_rate
#         })


    # グループ分け
    over_10 = []
    over_20 = []
    for record in records:
        if record["absence_rate"] > 0.2:
            over_20.append(record)
        elif record["absence_rate"] > 0.1:
            over_10.append(record)

    # レポート作成
    wb_out = Workbook()
    date_str = datetime.now().strftime("%Y年%m月%d日")

    def create_sheet(sheet_name, data):
        ws = wb_out.create_sheet(title=sheet_name)
        ws["A1"] = "出席状況レポート"
        ws["A2"] = f"報告日：{date_str}"
        ws.append(["", "", "", "", ""])
        ws.append(["生徒名", "科目", "欠席数", "授業数", "欠席率(%)"])

        for record in data:
            ws.append([
                record["student_name"],
                record["subject"],
                record["absence_count"],
                record["total_classes"],
                round(record["absence_rate"] * 100, 2)
            ])
        # セル整形
        for row in ws.iter_rows(min_row=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    create_sheet("欠席が多い生徒", over_10)
    create_sheet("欠席が基準を超えた生徒", over_20)

    # デフォルトシート削除
    del wb_out["Sheet"]

    wb_out.save(output_file)
    print(f"{output_file} を作成しました。")

# 実行
input_file = r"C:\projects\rpa2025\lesson09\科目別出席簿サンプル.xlsx"
output_file = "レポート.xlsx"
if not os.path.exists(input_file):
    print("指定されたファイルが見つかりません！:", input_file)
else:
    calc_absence_report(input_file, output_file)

# calc_absence_report(input_file, output_file)
