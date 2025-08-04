import requests
from openpyxl import Workbook
from collections import defaultdict

# APIから売上データを取得
url = "https://sales-api.yoshinori-wakai.workers.dev?start=2025-07-01&end=2025-07-31"
response = requests.get(url)
data = response.json()
print(type(data))
print(data[:1])

# 顧客ごとに売上をまとめる
customers = defaultdict(list)
for record in data:
    customer_name = record["customer_name"]  # ← 修正済み
    for item in record["items"]:
        customers[customer_name].append({
            "product": item["product_name"],
            "quantity": item["quantity"],
            "unit_price": item["unit_price"],
            "date": record["date"]
        })

# Excel作成
wb = Workbook()
ws = wb.active
ws.title = "請求書"

# ヘッダー
headers = ["顧客名", "商品名", "数量", "単価", "合計金額", "日付"]
ws.append(headers)

# データ書き込み
for customer, records in customers.items():
    for r in records:
        row = [
            customer,
            r["product"],
            r["quantity"],
            r["unit_price"],
            r["quantity"] * r["unit_price"],
            r["date"]
        ]
        ws.append(row)

# 保存
wb.save("請求書一覧.xlsx")
print("請求書を保存しました ✅")

# import requests
# from openpyxl import Workbook
# from collections import defaultdict

# # APIから売上データを取得
# url = "https://sales-api.yoshinori-wakai.workers.dev?start=2025-07-01&end=2025-07-31"
# response = requests.get(url)
# data = response.json()  # JSON形式のリスト
# print(type(data))
# print(data[:1])  # 最初の1件だけ表示して構造を確認


# # 顧客ごとに売上をまとめる
# customers = defaultdict(list)
# for record in data:
#     customer_name = record["customer_name"]  # 顧客名（仮）
#     customers[customer_name].append(record)

# # Excelブック作成
# wb = Workbook()
# ws = wb.active
# ws.title = "請求書"

# # ヘッダーを書き込む
# headers = ["顧客名", "商品名", "数量", "単価", "合計金額", "日付"]
# ws.append(headers)

# # データを書き込む
# for customer, records in customers.items():
#     for r in records:
#         row = [
#             customer_name,
#             r.get("product"),
#             r.get("quantity"),
#             r.get("unit_price"),
#             r.get("quantity") * r.get("unit_price"),
#             r.get("date")
#         ]
#         ws.append(row)

# # Excelとして保存
# wb.save("請求書一覧.xlsx")
# print("請求書を保存しました ✅")

# template_file = 'template_11.xlsx' #ファイル名を指定
# save_file = 'claim.xlsx' # 保存する名前を指定

# # 指定したURLからデータを取得したい

# customer_name =  ''# 顧客名
# date = ''          # 請求日
# total_amount = ''  # 合計金額
#                    # 品目 配列化した物は一応コメント化
# # items = [
# #     ['product_name','unit_price','quantity','total_price' ]
# # ]

# product_name = ''  # 商品名
# unit_price = ''    # 商品単価
# quantuty = ''      # 個数
# total_price = ''   # 小径





# import openpyxl as excel
# book = excel.load_workbook(template_file)
# sheet = book.active
# sheet["B6"] = customer_name  # sheetに読み込む
# # for i, it in enumerate(items):
# #     summary, count, price = it



# sheet["I33"] = total_amount 
# #sheet[""]
