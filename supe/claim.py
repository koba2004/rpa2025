import requests
from openpyxl import Workbook
from openpyxl.styles import Font
# セルに色を付けたいため「openpyxl.styles」を追加
from openpyxl.styles import Font, PatternFill

# 色の定義
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")     # 請求月・請求日（薄めの青）
light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # 請求額関係（もっと薄い水色）
header_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")     # 明細ヘッダー（薄い緑系などにしてもOK）
fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")


API_URL = "https://sales-api.yoshinori-wakai.workers.dev/?start=2025-07-01&end=2025-07-31"
CORPORATE_NAME = "株式会社APT"
BILLING_DATE = "2025年7月31日"
PREVIOUS_BILLING_AMOUNT = 0           #前回の請求額(今回は0円)
CURRENT_PAYMENT_AMOUNT = 0            #当月支払額(今回は0円)

def fetch_sales_data():
    response = requests.get(API_URL)  # requestsライブラリを使用して指定したAPIのURLからデータを取るようにリクエストを送る。
    response.raise_for_status()       # エラーを発生させる命令
    return response.json()            # APIから帰ってきたデータをJSON形式で返す

def create_invoice_for_customer(customer_name, details):
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"

    title_font = Font(size=20, bold=True)
    header_font = Font(bold=True)

    ws["A1"] = "御請求書"
    ws["A1"].font = title_font

    ws["A3"] = "請求月: 2025年7月分"
    ws["D3"] = f"請求日: {BILLING_DATE}"
    ws["A3"].fill = blue_fill
    ws["D3"].fill = blue_fill

    ws["A5"] = "御中"
    ws["B5"] = customer_name
    ws["A6"] = "自社名"
    ws["B6"] = CORPORATE_NAME

    ws["A8"] = "前回請求額"
    ws["B8"] = f"{PREVIOUS_BILLING_AMOUNT}円"
    ws["A8"].fill = light_blue_fill
    ws["B8"].fill = light_blue_fill
  
    ws["A9"] = "当月お支払い額"
    ws["B9"] = f"{CURRENT_PAYMENT_AMOUNT}円"
    ws["A9"].fill = light_blue_fill
    ws["B9"].fill = light_blue_fill

    total_amount = sum(item["total_price"] for item in details)
    ws["A10"] = "当月お買い上げ金額"
    ws["B10"] = f"{total_amount}円"
    ws["A10"].fill = light_blue_fill
    ws["B10"].fill = light_blue_fill

    ws["A11"] = "ご請求額"
    ws["B11"] = f"{total_amount}円"
    ws["A11"].fill = light_blue_fill
    ws["B11"].fill = light_blue_fill

    ws["A13"] = "お買い上げ明細"
    ws["A13"].font = header_font
    ws["A13"].font = header_font
    ws["A13"].fill = header_fill

    # 請求書のヘッダーに書き込むための処理
    headers = ["お買い上げ日", "商品コード", "商品名", "単価", "数量", "金額"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=14, column=col_num)
        cell.value = header      # 文字列の設定
        cell.font = header_font  # フォントスタイルの設定
        cell.fill = header_fill  # セルの色の設定

    # 明細を購入日、商品コードでソート（並び替え）
    #  今回は買い上げ日の昇順、商品コードの昇順で並び替え（ソート）をする指定があるため
    details_sorted = sorted(details, key=lambda x: (x["date"], x["product_id"]))

    for i, item in enumerate(details_sorted, start=15):
        ws.cell(row=i, column=1, value=item["date"])
        ws.cell(row=i, column=2, value=item["product_id"])
        ws.cell(row=i, column=3, value=item["product_name"])
        ws.cell(row=i, column=4, value=item["unit_price"])
        ws.cell(row=i, column=5, value=item["quantity"])
        ws.cell(row=i, column=6, value=item["total_price"])

    # ここではエクセルの横幅の調整を行っている。
    column_widths = [15, 15, 25, 10, 10, 15]
    for i, width in enumerate(column_widths, start=1):
        # chr() -> 文字コード(ASCllコード)から文字に変換する関数
        #  chr(64 + i) は、1 → A、2 → B、3 → C... としてExcelの列名（A〜F）を動的に作っている
        ws.column_dimensions[chr(64 + i)].width = width

       # セルの背景色をA1〜F30に設定（とても薄い緑）
    for row in ws.iter_rows(min_row=1, max_row=46, min_col=1, max_col=6):
        for cell in row:
            cell.fill = fill

    # エクセルファイルの出力
    filename = f"請求書_{customer_name}.xlsx"
    wb.save(filename)
    print(f"{filename} を作成しました。")

def main():
    transactions = fetch_sales_data() # fetch_sales_data()から取得した「取引データのリスト」

    # 顧客ごとに取引をまとめる dict[顧客名] = list[取引]
    #  (一人の顧客でも複数の購入がある)
    customers_transactions = {}  # 空の辞書を作成(ここに一人一人の取引リストを格納)
    for trans in transactions:
        customers_transactions.setdefault(trans["customer_name"], []).append(trans)
        # setdefault()・・・キーに顧客名のキーが辞書にまだないとき、空のリストを作成
        # .append(trans)で顧客リストに取引を追加

    # 顧客ごとに明細をフラット化
    for customer_name, trans_list in customers_transactions.items():
        details = []
        for trans in trans_list:
            for item in trans["items"]:
                # 明細に購入日と商品コード等をセット
                details.append({
                    "date": trans["date"],
                    "product_id": item["product_id"],
                    "product_name": item["product_name"],
                    "unit_price": item["unit_price"],
                    "quantity": item["quantity"],
                    "total_price": item["total_price"],
                })

        create_invoice_for_customer(customer_name, details)

# Pythonで書いたファイルが直接実行されたときにだけ、特定の処理を実行させる処理
if __name__ == "__main__":
    main()
