import openpyxl as excel
book = excel.load_workbook("all-customer.xlsx")
sheet = book["名簿"]
customers = [["名前","住所","購入プラン"]]
for row in sheet.iter_rows(min_row=3):
    values = [v.value for v in row]
    if values[0] is None: break
    area = values[1]
    if area == "横浜市" or area =="名古屋市":
        customers.append(values)
        print(values)


new_book = excel.Workbook()
new_sheet = new_book.active
new_sheet["A1"] = "横浜と名古屋の顧客名簿"

for row, row_val in enumerate(customers):
    for col, val in enumerate(row_val):
        c = new_sheet.cell(2+row, 1+col)
        c.value = val

new_book.save("yokohama_nagoya.xlsx""")