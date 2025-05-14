import openpyxl as excel
import random      # ランダム変数をインポート

atari = random.randint(1,100)

book = excel.Workbook()
book.active["B2"] = "あたりが書かれたシートを探そう"

for i in range(1,101):
    sname = str(i) + "番"
    sheet = book.create_sheet(title=sname)
    word = "はずれ"
    if i == atari:word = "あたり"
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y+1, x+1)
            c.value = word



book.save("game100.xlsx")
print("ok, atari=", atari)