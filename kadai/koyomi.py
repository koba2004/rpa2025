import openpyxl as excel  # openpyxlを取り込む

wareki_table = [
    {"name":"明治","start":1868, "end":1912},
    {"name":"大正","start":1912, "end":1926},
    {"name":"昭和","start":1926, "end":1989},
    {"name":"平成","start":1989, "end":2019},
    {"name":"令和","start":2019, "end":9999}
]

def seireki_wareki(year):
    for w in wareki_table:
        if w["start"] <= year < w["end"]:
            y = str(year - w["start"] + 1) + "年"
            if y == "1年": y = "元年"
            return w["name"] + y               # 和暦の更新
    return"不明"
    

book = excel.Workbook()   # 新規のワークブックの作成
sheet = book.active       # アクティブなワークシートを得る

sheet["A1"] = "和暦"
sheet["B1"] = "西暦"

start_y = 1930    # スタートを指定
for i in range(100): # 範囲を指定
    sei = start_y + i  # 上から順にたどっていく
    wa = seireki_wareki(sei) 

    sheet.cell(row=(2+i), column=1, value=str(sei)+"年") # 西暦を表示
    sheet.cell(row=(2+i), column=2, value=wa)            # 和暦を表示
    print(sei, "=", wa)


book.save("koyomi.xlsx")    # ファイルを保存